"""
Survey Image Processor with Hybrid AI Approach

Pipeline:
1. Azure AI Content Understanding - Custom analyzer extracts form structure
2. GPT-4o Vision - Validates extraction and flags low-confidence items
3. Excel Output - Structured workbook with confidence scores and review flags

Usage:
    python process_surveys.py --input surveys_sample/ --analyzer your-analyzer-id --output results.xlsx
"""

import os
import argparse
import requests
import time
import json
import pandas as pd
import base64
from pathlib import Path
from dotenv import load_dotenv
from typing import Dict, List, Tuple
from openai import AzureOpenAI
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Load environment variables
load_dotenv()

# Configuration
CONTENT_UNDERSTANDING_ENDPOINT = os.getenv("CONTENT_UNDERSTANDING_ENDPOINT")
CONTENT_UNDERSTANDING_KEY = os.getenv("CONTENT_UNDERSTANDING_KEY")
AZURE_OPENAI_ENDPOINT = os.getenv("AZURE_OPENAI_ENDPOINT")
AZURE_OPENAI_KEY = os.getenv("AZURE_OPENAI_KEY")
AZURE_OPENAI_DEPLOYMENT = os.getenv("AZURE_OPENAI_DEPLOYMENT", "gpt-4o")
AZURE_OPENAI_API_VERSION = os.getenv("AZURE_OPENAI_API_VERSION", "2024-05-01-preview")

# Confidence thresholds
LOW_CONFIDENCE_THRESHOLD = 0.70  # Re-extract if below 70%
REVIEW_FLAG_THRESHOLD = 0.85     # Flag for review if below 85%


def encode_image_to_base64(image_path: str) -> str:
    """Convert image to base64 for GPT-4o Vision"""
    with open(image_path, "rb") as f:
        return base64.b64encode(f.read()).decode('utf-8')


def analyze_with_content_understanding(image_path: str, analyzer_id: str) -> Tuple[dict, dict]:
    """
    Step 1: Extract data using Content Understanding custom analyzer

    Returns:
        - result: Extracted fields from Content Understanding
        - confidence_scores: Confidence score for each field
    """
    print(f"  [1/2] Analyzing with Content Understanding (analyzer: {analyzer_id})...")

    analyze_url = f"{CONTENT_UNDERSTANDING_ENDPOINT}/contentunderstanding/analyzers/{analyzer_id}:analyzeBinary?api-version=2025-11-01"

    headers = {
        "Ocp-Apim-Subscription-Key": CONTENT_UNDERSTANDING_KEY,
        "Content-Type": "image/jpeg"
    }

    with open(image_path, "rb") as f:
        image_data = f.read()

    # Submit analysis job
    response = requests.post(analyze_url, headers=headers, data=image_data)

    if response.status_code != 202:
        print(f"    Error: {response.status_code} - {response.text}")
        return None, {}

    result = response.json()
    result_id = result.get("id")

    # Poll for results
    result_url = f"{CONTENT_UNDERSTANDING_ENDPOINT}/contentunderstanding/analyzerResults/{result_id}?api-version=2025-11-01"

    while True:
        time.sleep(2)
        result_response = requests.get(result_url, headers={"Ocp-Apim-Subscription-Key": CONTENT_UNDERSTANDING_KEY})
        result_data = result_response.json()

        status = result_data.get("status")

        if status == "Succeeded":
            result = result_data.get("result")
            confidence_scores = extract_confidence_scores(result)
            return result, confidence_scores
        elif status == "Failed":
            print(f"    Analysis failed: {result_data.get('error', {}).get('message', 'Unknown error')}")
            return None, {}


def extract_confidence_scores(result: dict) -> dict:
    """Extract confidence scores from Content Understanding result"""
    if not result or "contents" not in result:
        return {}

    contents = result.get("contents", [])
    if not contents:
        return {}

    fields = contents[0].get("fields", {})

    confidences = {}
    for field_name, field_data in fields.items():
        if field_data and "confidence" in field_data:
            confidences[field_name] = field_data["confidence"]

    return confidences


def validate_with_gpt4_vision(image_path: str, extracted_data: dict, low_confidence_fields: List[str]) -> Tuple[str, str, List[str]]:
    """
    Step 2: Use GPT-4o Vision to validate extraction and flag issues

    Returns:
        - confidence: "High", "Medium", or "Low"
        - review_flag: Description of issues found, or "OK"
        - issues_found: List of specific issues
    """
    print(f"  [2/2] Validating with GPT-4o Vision...")

    if not AZURE_OPENAI_ENDPOINT or not AZURE_OPENAI_KEY:
        print("    Warning: Azure OpenAI not configured, skipping validation")
        return "Medium", "Validation skipped (no Azure OpenAI configured)", []

    # Initialize Azure OpenAI client
    client = AzureOpenAI(
        api_key=AZURE_OPENAI_KEY,
        api_version=AZURE_OPENAI_API_VERSION,
        azure_endpoint=AZURE_OPENAI_ENDPOINT
    )

    # Encode image
    base64_image = encode_image_to_base64(image_path)

    # Create summary of what Content Understanding found
    extraction_summary = json.dumps(extracted_data, indent=2)

    prompt = f"""You are validating survey data extraction. Compare the extracted data with the actual image.

EXTRACTED DATA:
{extraction_summary}

LOW CONFIDENCE FIELDS (need extra attention):
{', '.join(low_confidence_fields) if low_confidence_fields else 'None'}

REVIEW CHECKLIST:
1. Checkboxes: Are all visibly marked boxes captured? Any circles instead of checkmarks?
2. Text fields: Is handwriting correctly extracted?
3. Missing data: Any fields that should have data but don't?

Return JSON:
{{
  "confidence": "High" | "Medium" | "Low",
  "issues_found": ["issue 1", "issue 2"],
  "review_needed": true | false,
  "summary": "Brief assessment"
}}

CONFIDENCE LEVELS:
- High: Extraction looks accurate
- Medium: Minor issues or ambiguous marks
- Low: Major issues or missing data
"""

    response = client.chat.completions.create(
        model=AZURE_OPENAI_DEPLOYMENT,
        messages=[
            {
                "role": "system",
                "content": "You are a precise quality control reviewer for survey data extraction."
            },
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": prompt},
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:image/jpeg;base64,{base64_image}",
                            "detail": "high"
                        }
                    }
                ]
            }
        ],
        temperature=0,
        max_tokens=500,
        response_format={"type": "json_object"}
    )

    review_result = json.loads(response.choices[0].message.content)

    confidence = review_result.get("confidence", "Medium")
    issues = review_result.get("issues_found", [])
    review_flag = "; ".join(issues) if issues else "OK"

    return confidence, review_flag, issues


def process_survey_image(image_path: Path, analyzer_id: str) -> dict:
    """
    Process a single survey image through the hybrid pipeline

    Returns:
        Dictionary with extracted data, confidence scores, and flags
    """
    print(f"\nProcessing: {image_path.name}")

    try:
        # Step 1: Content Understanding extraction
        result, confidence_scores = analyze_with_content_understanding(str(image_path), analyzer_id)

        if not result:
            return {
                "image": image_path.name,
                "error": "Content Understanding analysis failed",
                "confidence": "Error"
            }

        # Extract fields from Content Understanding result
        extracted_data = {}
        if result.get("contents"):
            fields = result["contents"][0].get("fields", {})
            for field_name, field_data in fields.items():
                if field_data:
                    value = field_data.get("valueString") or field_data.get("valueSelectionMark")
                    extracted_data[field_name] = value

        # Step 2: Identify low-confidence fields
        low_confidence_fields = [
            field for field, score in confidence_scores.items()
            if score < LOW_CONFIDENCE_THRESHOLD
        ]

        # Step 3: Validate with GPT-4o Vision
        confidence, review_flag, issues = validate_with_gpt4_vision(
            str(image_path),
            extracted_data,
            low_confidence_fields
        )

        # Combine everything into result
        result_data = {
            "image": image_path.name,
            "confidence": confidence,
            "review_flag": review_flag,
            **extracted_data,
            **{f"{k}_confidence": f"{v:.0%}" for k, v in confidence_scores.items()}
        }

        # Print status
        status_icon = "✅" if confidence == "High" else "⚠️" if confidence == "Medium" else "❌"
        print(f"  {status_icon} Confidence: {confidence}")
        if review_flag != "OK":
            print(f"  ⚠️  Issues: {review_flag}")

        return result_data

    except Exception as e:
        print(f"  ❌ Error: {str(e)}")
        return {
            "image": image_path.name,
            "error": str(e),
            "confidence": "Error"
        }


def process_surveys_batch(input_dir: Path, analyzer_id: str, output_file: Path):
    """
    Process all survey images in a directory
    """
    print("="*60)
    print("SURVEY IMAGE PROCESSOR: Hybrid AI Pipeline")
    print("="*60)
    print(f"Input: {input_dir}")
    print(f"Analyzer: {analyzer_id}")
    print(f"Output: {output_file}")
    print()

    # Get all image files
    image_files = []
    for ext in ['*.jpg', '*.JPG', '*.jpeg', '*.JPEG', '*.png', '*.PNG']:
        image_files.extend(input_dir.glob(ext))

    image_files = sorted(set(image_files))

    if not image_files:
        print(f"No images found in {input_dir}")
        return

    print(f"Found {len(image_files)} images to process\n")

    # Process each image
    results = []
    for idx, image_path in enumerate(image_files, 1):
        print(f"[{idx}/{len(image_files)}]", end=" ")
        result = process_survey_image(image_path, analyzer_id)
        results.append(result)

    # Create DataFrame and save to Excel
    df = pd.DataFrame(results)
    df.to_excel(output_file, index=False)

    # Apply color coding to flagged items
    print("\nApplying visual formatting...")
    wb = load_workbook(output_file)
    ws = wb.active

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Highlight rows where confidence is not High
    for row_idx in range(2, ws.max_row + 1):
        confidence_cell = ws[f"C{row_idx}"]  # Assuming "confidence" is column C
        if confidence_cell.value and confidence_cell.value != "High":
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = yellow_fill

    wb.save(output_file)

    # Print summary
    print("\n" + "="*60)
    print("PROCESSING COMPLETE!")
    print("="*60)
    print(f"Total images: {len(image_files)}")
    print(f"High confidence: {len([r for r in results if r.get('confidence') == 'High'])}")
    print(f"Medium confidence: {len([r for r in results if r.get('confidence') == 'Medium'])}")
    print(f"Low confidence: {len([r for r in results if r.get('confidence') == 'Low'])}")
    print(f"Errors: {len([r for r in results if r.get('confidence') == 'Error'])}")
    print()
    print(f"✅ Results saved to: {output_file}")
    print()
    print("NEXT STEPS:")
    print("1. Open the Excel file")
    print("2. Review yellow-highlighted rows (medium/low confidence)")
    print("3. Check 'review_flag' column for specific issues")


def main():
    parser = argparse.ArgumentParser(description="Process survey images using hybrid AI pipeline")
    parser.add_argument("--input", required=True, help="Input directory containing survey images")
    parser.add_argument("--analyzer", required=True, help="Content Understanding analyzer ID")
    parser.add_argument("--output", default="survey_results.xlsx", help="Output Excel file")

    args = parser.parse_args()

    input_dir = Path(args.input)
    output_file = Path(args.output)

    if not input_dir.exists():
        print(f"Error: Input directory not found: {input_dir}")
        return

    if not CONTENT_UNDERSTANDING_ENDPOINT or not CONTENT_UNDERSTANDING_KEY:
        print("Error: Content Understanding credentials not configured in .env file")
        return

    process_surveys_batch(input_dir, args.analyzer, output_file)


if __name__ == "__main__":
    main()
